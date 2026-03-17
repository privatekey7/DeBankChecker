"""
Экспорт результатов в Excel
"""

import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from debank_checker.export.common import filter_nft, filter_protocols, filter_tokens
from debank_checker.export.config import ExportConfig


def export_to_excel(
    results: list[dict],
    output_dir: Path | str = ".",
    config: ExportConfig | None = None,
) -> Path:
    """
    Экспорт в Excel по ExportConfig.
    Если config=None — экспорт всего (как раньше).
    """
    config = config or ExportConfig(summary=True, tokens=True, nft=True, protocols=True)
    output_dir = Path(output_dir)
    suffix = config.get_filename_suffix()
    filename = f"debank_{suffix}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = output_dir / filename

    wb = Workbook()

    if config.total_only:
        _add_total_only_sheet(wb, results)
    else:
        default_sheet = wb.active
        if config.summary:
            _style_summary_sheet(wb, results)
        elif not any([config.tokens, config.protocols, config.nft]):
            _style_summary_sheet(wb, results)
        if config.tokens:
            _add_tokens_sheet(wb, results, config.token_filter)
        if config.protocols:
            _add_protocols_sheet(wb, results, config.protocol_filter)
        if config.nft:
            _add_nft_sheet(wb, results, config.nft_filter)
        if not config.summary and any([config.tokens, config.protocols, config.nft]) and default_sheet in wb.worksheets:
            wb.remove(default_sheet)

    wb.save(filepath)
    return filepath


def _add_total_only_sheet(wb: Workbook, results: list[dict]) -> None:
    """Лист только с общим балансом."""
    ws = wb.active
    ws.title = "Total"
    total = sum(r["total_usd"] for r in results if r and r.get("status") == "OK")
    ok = sum(1 for r in results if r and r.get("status") == "OK")
    ws.cell(row=1, column=1, value="Суммарный баланс").font = Font(bold=True)
    ws.cell(row=1, column=2, value=total).number_format = "#,##0.00"
    ws.cell(row=2, column=1, value="Успешно")
    ws.cell(row=2, column=2, value=f"{ok}/{len(results)}")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18


def _style_summary_sheet(wb: Workbook, results: list[dict]) -> None:
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A73E8")
    ok_fill = PatternFill("solid", fgColor="E6F4EA")
    err_fill = PatternFill("solid", fgColor="FCE8E6")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "Адрес", "USD баланс", "Chains", "Статус"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = max(15, len(h) + 2)

    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18

    for ri, r in enumerate(results, 2):
        row_fill = ok_fill if r["status"] == "OK" else err_fill
        values = [
            ri - 1,
            r["address"],
            r["total_usd"] if r["status"] == "OK" else None,
            r.get("chains", ""),
            r["status"],
        ]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = row_fill
            c.border = border
            c.alignment = center if ci in (1, 3, 5) else left
            if ci == 3 and isinstance(val, (int, float)):
                c.number_format = "#,##0.00"

    total_sum = sum(x["total_usd"] for x in results if x["status"] == "OK")
    ok_count = sum(1 for x in results if x["status"] == "OK")
    summary_row = len(results) + 2
    ws.cell(row=summary_row, column=1, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"{ok_count}/{len(results)} успешно")
    sc = ws.cell(row=summary_row, column=3, value=total_sum)
    sc.font = Font(bold=True)
    sc.number_format = "#,##0.00"

    ws.freeze_panes = "A2"


def _add_tokens_sheet(wb: Workbook, results: list[dict], token_filter: str | None = None) -> None:
    ws = wb.create_sheet("Tokens", 1)
    headers = ["Адрес", "Symbol", "Chain", "Amount", "Price", "Value USD"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(ci)].width = 15
    ws.column_dimensions["A"].width = 45

    row = 2
    for r in results:
        if r["status"] != "OK":
            continue
        tokens = filter_tokens(r.get("tokens_data", []), token_filter)
        for t in tokens:
            ws.cell(row=row, column=1, value=r["address"])
            ws.cell(row=row, column=2, value=t.get("symbol", ""))
            ws.cell(row=row, column=3, value=t.get("chain", ""))
            ws.cell(row=row, column=4, value=t.get("amount", 0))
            ws.cell(row=row, column=5, value=t.get("price", 0))
            ws.cell(row=row, column=6, value=t.get("value", 0))
            row += 1

    if token_filter and row == 2:
        ws.cell(row=2, column=1, value=f"Нет совпадений по фильтру {token_filter}. Проверьте symbol и chain.")


def _add_protocols_sheet(wb: Workbook, results: list[dict], protocol_filter: str | None = None) -> None:
    ws = wb.create_sheet("Protocols", 2)
    headers = ["Адрес", "Протокол", "Chain", "Value USD"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.column_dimensions["A"].width = 45

    row = 2
    for r in results:
        if r["status"] != "OK":
            continue
        protocols = filter_protocols(r.get("protocols_data", []), protocol_filter)
        for p in protocols:
            ws.cell(row=row, column=1, value=r["address"])
            ws.cell(row=row, column=2, value=p.get("name", ""))
            ws.cell(row=row, column=3, value=p.get("chain", ""))
            ws.cell(row=row, column=4, value=p.get("value", 0))
            row += 1

    if protocol_filter and row == 2:
        ws.cell(row=2, column=1, value=f"Нет совпадений по фильтру {protocol_filter}. Проверьте protocol и chain.")


def _add_nft_sheet(wb: Workbook, results: list[dict], nft_filter: str | None = None) -> None:
    ws = wb.create_sheet("NFT", 3)
    headers = ["Адрес", "Collection", "Chain", "Amount"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.column_dimensions["A"].width = 45

    row = 2
    for r in results:
        if r["status"] != "OK":
            continue
        nfts = filter_nft(r.get("nft_data", []), nft_filter)
        for n in nfts:
            ws.cell(row=row, column=1, value=r["address"])
            ws.cell(row=row, column=2, value=n.get("name", ""))
            ws.cell(row=row, column=3, value=n.get("chain", ""))
            ws.cell(row=row, column=4, value=n.get("amount", 0))
            row += 1

    if nft_filter and row == 2:
        ws.cell(row=2, column=1, value=f"Нет совпадений по фильтру {nft_filter}. Проверьте collection и chain.")
